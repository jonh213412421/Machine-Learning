import os
import sys
import time
import google.generativeai as genai
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from tkinter import filedialog
from PIL import Image

# Local do Webdriver. Automatizar o download dele futuramente?
WEBDRIVER = os.path.join(os.getcwd(), "webdriver", "msedgedriver.exe")
VERIFICACAO_TRIPLA = True


def iniciar() -> None:
    try:
        if not os.path.exists(os.path.join(os.getcwd(), "webdriver", "msedgedriver.exe")):
            print("Erro: Edge Driver não foi localizado. Ele deve estar localizado na pasta 'webdriver'")
        if not os.path.exists(os.path.join(os.getcwd(), "bancos de dados")):
            os.mkdir(os.path.join(os.getcwd(), "bancos de dados"))
    except Exception as e:
        print(f"Erro na inicializacao {e}")

    try:
        if not os.path.exists(os.path.join(os.getcwd(), "config")):
            key = input("qual é a chave API?")
            # Configuração da chave de API da Gemini. Colocar em variável de ambiente futuramente.
            genai.configure(api_key=key)
            salvar = input("deseja salvar a chave? (s/n)")
            if salvar == "s":
                os.mkdir(os.path.join(os.getcwd(), "config"))
                if not os.path.isfile(os.path.join(os.getcwd(), "config", "config.ini")):
                    with open(os.path.join(os.getcwd(), "config", "config.ini"), "w") as f:
                        f.write(key)
        else:
            with open(os.path.join(os.getcwd(), "config", "config.ini"), "r") as f:
                key = f.read()
            # Configuração da chave de API da Gemini. Colocar em variável de ambiente futuramente.
            genai.configure(api_key=key)
    except Exception as e:
        print(f"Erro ao salvar chave: {e}")


# Pega os dados do site usando selenium headless. É possível alterar a posição da janela usando o argumento comentado.
def obter_dados_texto(url_site: str) -> str:
    options = EdgeOptions()
    options.add_argument("--headless")  # Modo headless
    # options.add_argument("--window-position=3000,0")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")

    # Cria o objeto Service com o caminho do driver. O local do executável está declarado na variável global "WEBDRIVER"
    service = EdgeService(executable_path=WEBDRIVER)

    # Inicia o driver com o Edge headless
    driver = webdriver.Edge(
        service=service,
        options=options
    )
    try:
        driver.get(url_site)
        time.sleep(1)
        return driver.page_source
    except Exception as e:
        return f"erro ao obter código da página: {e}"


#
def perguntar_a_ia_texto(texto: str, vars: str) -> str:
    # Melhorar, tornar mais robusto. A ideia é pegar três saídas diferentes, comparar e retornar o mais frequente. No entanto, usar um sorted()
    # melhora a robustez. Pode ser facilmente expandido para verificar mais do que três vezes.
    def verificar_consistencia(respostas: list) -> str:
        try:
            # atribui cada saída gerada pela IA para um vetor diferente.
            auxa = respostas[0]
            auxa = auxa.splitlines()[1:-1]
            auxb = respostas[1]
            auxb = auxb.splitlines()[1:-1]
            auxc = respostas[2]
            auxc = auxc.splitlines()[1:-1]
            # Comparar aux e ver qual se repete mais. Ordena as listas
            print(f"{sorted(set(auxa))} -> {sorted(set(auxb))} -> {sorted(set(auxc))}")
            if sorted(set(auxa)) == sorted(set(auxb)):
                return "\n".join(auxa)
            if sorted(set(auxb)) == sorted(set(auxc)):
                return "\n".join(auxb)
            if sorted(set(auxa)) == sorted(set(auxc)):
                return "\n".join(auxc)
            else:
                op = input(
                    "Erro na consistência! Recomendado tentar novamente.\n1) continuar\n2) sair\n3) rodar novamente")
                if op == "1":
                    return auxa
                elif op == "2":
                    sys.exit()
                elif op == "3":
                    print("Tentando novamente...")
                    perguntar_a_ia_texto(texto, vars)
                else:
                    print("Opção inválida! Saindo do Programa...")
                    time.sleep(10)
                    sys.exit()
        except Exception as e:
            print(f"Erro na verificação de consistencia: {e}")

    vars_aux = vars.split("--")
    vars_aux = ", ".join(vars_aux)
    """Usa a IA Gemini para responder a uma pergunta baseada no texto fornecido."""
    model = genai.GenerativeModel('gemini-2.0-flash')

    # Prepara o prompt para a IA
    # É sempre importante explicar o máximo possível e colocar exemplos nos prompts
    prompt = f"""
    Analise a página seguir.
    com base nela, crie uma tabela e a preencha com os seguintes dados: {vars_aux}. Cada um desses dados deve representar
    uma coluna na tabela. os dados da tabela devem ser separados com uma barra '/'. Não coloque títulos nas colunas, isso é muito importante!
    Se atenha apenas aos dados fornecidos. Extraia todos os dados da página.
    Exemplo 1:
    dados: "idade", "altura", "tipo sanguíneo"
    resultado: 
    15/1.67/O+
    17/1.75/O-
    18/1.80/B+

    Exemplo 2:
    dados: "país", "PIB em bilhões de dólares", "índice Gini"
    resultado: 
    Brasil/10.000/0.706
    EUA/100.000/0.850
    Japão/50.000/0.900

    Exemplo 3:
    dados: "preço", "quantidade", "estoque"
    resultado: 
    5.50/7/500
    7.50/13/100
    9/4/25

    PÁGINA:
    {texto}
    """
    try:
        if VERIFICACAO_TRIPLA:
            resposta = []
            for i in range(0, 3):
                resposta.append(model.generate_content(prompt).text)
                print(resposta)

            lista_final = verificar_consistencia(resposta)
            return lista_final
        else:
            resposta = model.generate_content(prompt).text.splitlines()[1:-1]
            return resposta

    except Exception as e:
        return f"Erro ao gerar conteúdo com a IA: {e}"


def perguntar_a_ia_imagem(caminho_imagem: str, vars: str) -> str:
    # Melhorar, tornar mais robusto. A ideia é pegar três saídas diferentes, comparar e retornar o mais frequente. No entanto, usar um sorted()
    # melhora a robustez. Pode ser facilmente expandido para verificar mais do que três vezes.
    def verificar_consistencia(respostas: list) -> str:
        try:
            # atribui cada saída gerada pela IA para um vetor diferente.
            auxa = respostas[0]
            auxa = auxa.splitlines()[1:-1]
            auxb = respostas[1]
            auxb = auxb.splitlines()[1:-1]
            auxc = respostas[2]
            auxc = auxc.splitlines()[1:-1]
            # Comparar aux e ver qual se repete mais. Ordena as listas
            print(f"{sorted(set(auxa))} -> {sorted(set(auxb))} -> {sorted(set(auxc))}")
            if sorted(set(auxa)) == sorted(set(auxb)):
                return "\n".join(auxa)
            if sorted(set(auxb)) == sorted(set(auxc)):
                return "\n".join(auxb)
            if sorted(set(auxa)) == sorted(set(auxc)):
                return "\n".join(auxc)
            else:
                op = input(
                    "Erro na consistência! Recomendado tentar novamente.\n1) continuar\n2) sair\n3) rodar novamente")
                if op == "1":
                    return auxa
                elif op == "2":
                    sys.exit()
                elif op == "3":
                    print("Tentando novamente...")
                    perguntar_a_ia_imagem(caminho_imagem, vars)
                else:
                    print("Opção inválida! Saindo do Programa...")
                    time.sleep(10)
                    sys.exit()
        except Exception as e:
            print(f"Erro na verificação de consistencia: {e}")

    vars = vars.split("--")
    vars = ", ".join(vars)
    imagem = Image.open(caminho_imagem)
    """Usa a IA Gemini para responder a uma pergunta baseada no texto fornecido."""
    model = genai.GenerativeModel('gemini-2.5-flash-image-preview')

    # Prepara o prompt para a IA
    # É sempre importante explicar o máximo possível e colocar exemplos nos prompts
    prompt = f"""
    Analise a imagem a seguir. Use o reconhecimento de imagem.
    com base nela, crie uma tabela e a preencha com os seguintes dados: {vars}. Cada um desses dados deve representar
    uma coluna na tabela. os dados da tabela devem ser separados com uma barra '/'. Não coloque títulos nas colunas, isso é muito importante!
    Se atenha apenas aos dados fornecidos. Extraia todos os dados da página.
    Exemplo 1:
    dados: "idade", "altura", "tipo sanguíneo"
    resultado: 
    15/1.67/O+
    17/1.75/O-
    18/1.80/B+

    Exemplo 2:
    dados: "país", "PIB em bilhões de dólares", "índice Gini"
    resultado: 
    Brasil/10.000/0.706
    EUA/100.000/0.850
    Japão/50.000/0.900

    Exemplo 3:
    dados: "preço", "quantidade", "estoque"
    resultado: 
    5.50/7/500
    7.50/13/100
    9/4/25
    """

    try:
        resposta = model.generate_content([prompt, imagem])
        return resposta.text
    except Exception as e:
        return f"Erro ao gerar conteúdo com a IA: {e}"


# salva os dados em uma planilha
def salvar_planilha(vars, dados) -> None:
    try:
        # elimina os parêntesis no início e no final
        dados = dados.splitlines()
        print(dados)
        num_linhas = len(dados)
        # extrai os argumentos
        vars = vars.split("--")[1:]
        print(vars)
        num_cols = len(vars)
        print(f"\n{num_linhas} linhas ne planilha.")
        print(f"\n{num_cols} colunas na planilha.")
        wb = Workbook()
        ws = wb.active
        for i in range(num_cols):
            ws.cell(row=1, column=i + 1, value=vars[i])
            for linha in range(num_linhas):
                try:
                    ws.cell(row=linha + 2, column=i + 1, value=dados[linha].split("/")[i])
                except Exception as e:
                    print(e)
        # Salva a planilha com um nome único. Importante para não haver sobreposição de dados.
        nome_arquivo_saida = os.path.join(os.getcwd(), "bancos de dados",
                                          f"database-{time.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx")
        wb.save(os.path.join(nome_arquivo_saida))
        print(f"Dados salvos! Nome do arquivo: {nome_arquivo_saida}\n\n")
    except Exception as e:
        print(f"Erro: {e}")


# concatena planilhas para criar banco de dados únicos
def cat_data(caminhos: list) -> None:
    try:
        # o i elimina a primeira linha.
        i = 0
        wb_combinado = Workbook()
        ws = wb_combinado.active
        for caminho in caminhos:
            wb = load_workbook(caminho)
            sh = wb.active
            for linha in sh.iter_rows(values_only=True):
                if i > 0:
                    ws.append(linha)
                i += 1
            i = 0
        wb_combinado.save(os.path.join(os.getcwd(), "bancos de dados",
                                       f"database-combined-{time.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx"))
        print("Pronto! Dados concatenados!")
    except Exception as e:
        print(f"Erro na combinação das bases de dados: {e}")


if __name__ == "__main__":
    iniciar()
    # Loop infinito
    while True:
        try:
            op = input(
                "O que deseja fazer?\n1) Extrair dados de um site\n2) Extrair dados de uma imagem\n3) Concatenar dados\n4) Configurações\n")
            # Para extrair de um site
            if op == "1":
                try:
                    url_do_site = input("entre com o site que quer extrair informações: ")
                    vars = input("quais variáveis deseja extrair? separe-as com '--'.\n ")
                    if not url_do_site.startswith("http://") and not url_do_site.startswith("https://"):
                        url_do_site = "https://" + url_do_site
                    nome_do_arquivo_pdf = url_do_site.replace("http://", "")
                    nome_do_arquivo_pdf = nome_do_arquivo_pdf.replace("https://", "")
                    nome_do_arquivo_pdf = nome_do_arquivo_pdf.replace("www.", "")
                    nome_do_arquivo_pdf = nome_do_arquivo_pdf.split(".")[0]
                    nome_do_arquivo_pdf = nome_do_arquivo_pdf + ".pdf"

                    # 1. Converte o site para PDF
                    conteudo_do_site = obter_dados_texto(url_do_site)

                    # 3. Faz a pergunta à IA com base no texto extraído
                    dados = perguntar_a_ia_texto(conteudo_do_site, vars)
                    print("\nResposta da IA:\n")
                    print(dados)
                    salvar_planilha(vars, dados)
                except Exception as e:
                    print(f"Erro: {e}")

            elif op == "2":
                try:
                    caminho_imagem = filedialog.askopenfile(initialdir=os.getcwd(),
                                                            filetypes=[("Todos", "."), ("PDF", ".pdf"), ("JPG", ".jpg"),
                                                                       ("JPEG", ".jpeg"), ("PNG", ".png")],
                                                            title="Escolha o arquivo de imagem")
                    vars = input("quais variáveis deseja extrair? separe-as com '--'.\n ")
                    caminho_imagem = caminho_imagem.name
                    dados = perguntar_a_ia_imagem(caminho_imagem, vars)
                    print("\nResposta da IA:\n")
                    print(dados)
                    print(f"caminho da imagem: {caminho_imagem}")
                    salvar_planilha(vars, dados)
                except Exception as e:
                    print(f"Erro: {e}")

            elif op == "3":
                try:
                    caminhos = filedialog.askopenfilenames(initialdir=os.getcwd(), title="Escolha os arquivos")
                    if caminhos:
                        cat_data(caminhos)
                    else:
                        print("Nenhum arquivo foi selecionado.")
                except Exception as e:
                    print(f"Erro: {e}")

            elif op == "4":
                while True:
                    print(f"1- Redundância (executa 3 vezes para verificar consistência de output) - {VERIFICACAO_TRIPLA}\n2) Voltar")
                    alt = input("Digite a configuração que deseja alterar: ")
                    if alt == "1":
                        VERIFICACAO_TRIPLA = not VERIFICACAO_TRIPLA
                    if alt == "2":
                        print("\n")
                        break

            else:
                print("Operação Inválida! Tente novamente.")
        except Exception as e:
            print(f"Erro no loop principal: {e}")
